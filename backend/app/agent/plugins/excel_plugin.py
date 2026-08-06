import logging
from typing import Any

import openpyxl
from app.agent.base_plugin import PluginBase, PluginError, PluginOutput
from app.agent.registry import register_plugin
from app.services.artifact_service import ArtifactService
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, field_validator

logger = logging.getLogger(__name__)

MAX_EXCEL_ROWS = 100000


class ExcelPluginInput(BaseModel):
    title: str = Field(..., description="Workbook title or topic")
    sheet_name: str | None = Field("Analytics Data", description="Name of the excel worksheet")
    data: list[dict[str, Any]] | str | None = Field(None, description="Optional dataset rows. If omitted or using previous query result, omit or leave null.")

    @field_validator("data", mode="before")
    @classmethod
    def sanitize_data(cls, v: Any) -> list[dict[str, Any]] | None:
        if isinstance(v, str) or not isinstance(v, list):
            return None
        return v


@register_plugin
class ExcelPlugin(PluginBase):
    @property
    def name(self) -> str:
        return "excel"

    @property
    def description(self) -> str:
        return "Generates a styled Excel (.xlsx) workbook with formatted headers, column scaling, and bounds checking."

    @property
    def input_schema(self) -> type[BaseModel]:
        return ExcelPluginInput

    @property
    def can_consume(self) -> list[str]:
        return ["data"]

    @property
    def can_produce(self) -> list[str]:
        return ["file"]

    async def execute(
        self,
        params: dict[str, Any],
        context: dict[str, Any],
        progress_callback: Any | None = None
    ) -> PluginOutput:
        val_err = self.validate_params(params)
        if val_err:
            return PluginOutput(success=False, output_type="excel", error=val_err)

        title = params.get("title", "Analytics Report")
        sheet_name = params.get("sheet_name", "Data")
        dataset = params.get("data")
        if not isinstance(dataset, list):
            dataset = context.get("last_query_result") or []

        if not dataset:
            return PluginOutput(
                success=False,
                output_type="excel",
                error=PluginError(code="EMPTY_DATASET", message="No data available to export to Excel.", retryable=True)
            )

        # Enforce max row limit bound
        if len(dataset) > MAX_EXCEL_ROWS:
            dataset = dataset[:MAX_EXCEL_ROWS]

        if progress_callback:
            await progress_callback(f"Formatting Excel workbook: {title} ({len(dataset)} rows)...")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:30]

        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = title.upper()
        title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        headers = list(dataset[0].keys())
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        ws.row_dimensions[3].height = 25

        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        for row_idx, row_data in enumerate(dataset, 4):
            for col_idx, header in enumerate(headers, 1):
                val = row_data.get(header)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row >= 3 and cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        artifact_id, filepath = ArtifactService.create_artifact_file(".xlsx")
        wb.save(filepath)

        return PluginOutput(
            success=True,
            output_type="excel",
            result={"filename": f"{artifact_id}.xlsx", "rows_count": len(dataset)},
            artifact_id=artifact_id,
            artifact_url=f"/api/artifacts/{artifact_id}",
            metadata={"title": title, "rows": len(dataset)}
        )
